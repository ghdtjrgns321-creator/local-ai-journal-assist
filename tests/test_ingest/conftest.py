"""test_ingest 공용 fixture — 테스트용 임시 파일 생성."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq
import pytest


@pytest.fixture
def valid_xlsx(tmp_path: Path) -> Path:
    """최소 데이터가 포함된 정상 .xlsx 파일."""
    filepath = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["journal_id", "entry_date", "debit_amount"])
    ws.append([1, "2025-01-01", 10000])
    wb.save(filepath)
    return filepath


@pytest.fixture
def valid_csv(tmp_path: Path) -> Path:
    """UTF-8 인코딩 정상 .csv 파일."""
    filepath = tmp_path / "test.csv"
    filepath.write_text(
        "journal_id,entry_date,debit_amount\n1,2025-01-01,10000\n",
        encoding="utf-8",
    )
    return filepath


@pytest.fixture
def valid_csv_cp949(tmp_path: Path) -> Path:
    """CP949 인코딩 .csv 파일 — 인코딩 경고 테스트용."""
    filepath = tmp_path / "test_cp949.csv"
    filepath.write_text(
        "전표번호,일자,차변금액\n1,2025-01-01,10000\n",
        encoding="cp949",
    )
    return filepath


@pytest.fixture
def valid_tsv(tmp_path: Path) -> Path:
    """정상 .tsv 파일."""
    filepath = tmp_path / "test.tsv"
    filepath.write_text(
        "journal_id\tentry_date\tdebit_amount\n1\t2025-01-01\t10000\n",
        encoding="utf-8",
    )
    return filepath


@pytest.fixture
def valid_parquet(tmp_path: Path) -> Path:
    """정상 .parquet 파일."""
    filepath = tmp_path / "test.parquet"
    table = pa.table({"journal_id": [1], "amount": [10000.0]})
    pq.write_table(table, filepath)
    return filepath


@pytest.fixture
def xlsx_multi_sheet(tmp_path: Path) -> Path:
    """3개 시트(데이터2 + 빈1) .xlsx 파일."""
    filepath = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()

    # Sheet1 (활성 시트) — 데이터 있음
    ws1 = wb.active
    ws1.title = "매출"
    ws1.append(["id", "date", "amount"])
    ws1.append([1, "2025-01-01", 5000])

    # Sheet2 — 데이터 있음
    ws2 = wb.create_sheet("매입")
    ws2.append(["id", "date", "amount"])
    ws2.append([2, "2025-02-01", 3000])

    # Sheet3 — 빈 시트
    wb.create_sheet("빈시트")

    wb.save(filepath)
    return filepath


@pytest.fixture
def xlsx_with_merged_cells(tmp_path: Path) -> Path:
    """병합셀이 포함된 .xlsx 파일."""
    filepath = tmp_path / "merged.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    # 행1: 병합 헤더 (A1:B1 병합 = "회사정보")
    ws["A1"] = "회사정보"
    ws.merge_cells("A1:B1")
    ws["C1"] = "금액"

    # 행2: 실제 데이터
    ws["A2"] = "코드"
    ws["B2"] = "이름"
    ws["C2"] = 10000

    # 행3: 세로 병합 테스트 (A3:A4 병합 = "C001")
    ws["A3"] = "C001"
    ws.merge_cells("A3:A4")
    ws["B3"] = "본사"
    ws["C3"] = 20000
    ws["B4"] = "지사"
    ws["C4"] = 30000

    wb.save(filepath)
    return filepath


@pytest.fixture
def csv_with_bom(tmp_path: Path) -> Path:
    """BOM(UTF-8-SIG) 포함 .csv 파일."""
    filepath = tmp_path / "bom.csv"
    filepath.write_text(
        "id,name,amount\n1,테스트,10000\n",
        encoding="utf-8-sig",
    )
    return filepath


@pytest.fixture
def corrupted_xlsx(tmp_path: Path) -> Path:
    """확장자는 .xlsx이지만 내용이 랜덤 바이트인 손상 파일."""
    filepath = tmp_path / "corrupted.xlsx"
    filepath.write_bytes(b"\x00\x01\x02\x03NOTANEXCELFILE")
    return filepath


@pytest.fixture
def empty_file(tmp_path: Path) -> Path:
    """0바이트 빈 파일."""
    filepath = tmp_path / "empty.xlsx"
    filepath.write_bytes(b"")
    return filepath


@pytest.fixture
def pdf_file(tmp_path: Path) -> Path:
    """.pdf 확장자 파일 (내용 무관)."""
    filepath = tmp_path / "report.pdf"
    filepath.write_bytes(b"%PDF-1.4 dummy")
    return filepath


@pytest.fixture
def hwp_file(tmp_path: Path) -> Path:
    """.hwp 확장자 파일 (내용 무관)."""
    filepath = tmp_path / "report.hwp"
    filepath.write_bytes(b"\xd0\xcf\x11\xe0 dummy")
    return filepath


# ── 헤더 탐지(header_detector) 전용 fixture ──────────────────


@pytest.fixture
def hd_standard_xlsx(tmp_path: Path) -> Path:
    """표준 1행 헤더 — keywords.yaml 별칭 사용."""
    filepath = tmp_path / "hd_standard.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["전표번호", "전표일자", "계정코드", "차변금액", "대변금액", "적요"])
    ws.append([1, "2025-01-01", "1110", 10000, 0, "사무용품 구매"])
    ws.append([2, "2025-01-02", "2110", 0, 5000, "매출 입금"])
    wb.save(filepath)
    return filepath


@pytest.fixture
def hd_erp_style_xlsx(tmp_path: Path) -> Path:
    """ERP 스타일 — 1~2행 제목, 3행(idx=2) 헤더."""
    filepath = tmp_path / "hd_erp.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["[결산 보고서]", None, None, None])
    ws.append(["작성일: 2025-03-18", None, None, None])
    ws.append(["전표번호", "전표일자", "차변금액", "대변금액"])
    ws.append([1, "2025-01-01", 10000, 0])
    wb.save(filepath)
    return filepath


@pytest.fixture
def hd_merged_header_xlsx(tmp_path: Path) -> Path:
    """병합셀 상위 + 실제 헤더 2행(idx=1).

    pandas read_excel(header=None)은 병합셀의 첫 번째 셀에만 값을 넣고
    나머지는 NaN으로 처리 → 1행은 문자열 2개뿐(키워드 없음, confidence~0.2),
    2행은 키워드 5개 → row=1이 정확히 탐지되는 구조를 검증.
    """
    filepath = tmp_path / "hd_merged.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    # Row 1: 병합 그룹명
    ws["A1"] = "전표 정보"
    ws.merge_cells("A1:C1")
    ws["D1"] = "금액 정보"
    ws.merge_cells("D1:E1")
    # Row 2: 실제 헤더
    ws["A2"] = "전표번호"
    ws["B2"] = "전표일자"
    ws["C2"] = "계정코드"
    ws["D2"] = "차변금액"
    ws["E2"] = "대변금액"
    # Row 3: 데이터
    ws["A3"] = 1
    ws["B3"] = "2025-01-01"
    ws["C3"] = "1110"
    ws["D3"] = 10000
    ws["E3"] = 0
    wb.save(filepath)
    return filepath


@pytest.fixture
def hd_non_accounting_xlsx(tmp_path: Path) -> Path:
    """비회계 데이터 — 키워드 매칭 실패 기대."""
    filepath = tmp_path / "hd_non_acct.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["이름", "부서", "직급", "연락처"])
    ws.append(["홍길동", "경영지원", "과장", "010-1234-5678"])
    wb.save(filepath)
    return filepath


@pytest.fixture
def hd_dirty_columns_xlsx(tmp_path: Path) -> Path:
    """유효 키워드 + 대량 빈 컬럼 혼재."""
    filepath = tmp_path / "hd_dirty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    row = ["전표번호", "전표일자", "차변금액", "대변금액"] + [None] * 16
    ws.append(row)
    ws.append([1, "2025-01-01", 10000, 0] + [None] * 16)
    wb.save(filepath)
    return filepath


@pytest.fixture
def hd_trick_data_xlsx(tmp_path: Path) -> Path:
    """데이터 행에 키워드가 섞인 트릭 케이스."""
    filepath = tmp_path / "hd_trick.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["전표번호", "전표일자", "차변금액", "대변금액", "적요"])
    ws.append([1, "2025-01-01", 10000, 0, "전표일자 수정 요청"])
    ws.append([2, "2025-01-02", 0, 5000, "차변 금액 확인"])
    wb.save(filepath)
    return filepath


@pytest.fixture
def hd_mid_confidence_xlsx(tmp_path: Path) -> Path:
    """중간 신뢰도(0.3~0.7) 강제 — 키워드 1개만 매칭.

    keyword_score = 1/4 = 0.25, string_ratio = 1.0
    → confidence ≈ 0.25*0.8 + 1.0*0.2 = 0.4 (중간 구간)
    """
    filepath = tmp_path / "hd_mid.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["전표번호", "날짜", "금액", "구분"])
    ws.append([1, "2025-01-01", 10000, "매출"])
    wb.save(filepath)
    return filepath


@pytest.fixture
def hd_sap_style_xlsx(tmp_path: Path) -> Path:
    """SAP 스타일 영문/코드명 헤더."""
    filepath = tmp_path / "hd_sap.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["BELNR", "BUDAT", "HKONT", "Debit Amount", "Credit Amount"])
    ws.append(["5000001", "2025-01-01", "1110", 10000, 0])
    wb.save(filepath)
    return filepath


# ── 컬럼 매핑(column_mapper) 전용 fixture ──────────────────


@pytest.fixture
def cm_standard_columns() -> list[str]:
    """DataSynth 표준 스키마 컬럼명 — fast path 기대."""
    return [
        "document_id", "company_code", "fiscal_year",
        "posting_date", "document_date", "gl_account",
        "debit_amount", "credit_amount", "document_type",
        "created_by", "source", "line_text",
    ]


@pytest.fixture
def cm_korean_columns() -> list[str]:
    """한국어 별칭 — exact match 기대."""
    return [
        "전표번호", "회사코드", "회계연도",
        "전표일자", "증빙일자", "계정코드",
        "차변금액", "대변금액", "전표유형",
        "작성자", "입력구분", "적요",
    ]


@pytest.fixture
def cm_sap_columns() -> list[str]:
    """SAP ACDOCA 코드명 — exact match 기대."""
    return [
        "belnr", "bukrs", "gjahr",
        "budat", "bldat", "racct",
        "debit_amount", "credit_amount", "blart",
        "usnam", "source", "sgtxt",
    ]


@pytest.fixture
def cm_mixed_columns() -> list[str]:
    """한글 + 영문 + 비표준 혼합 — fuzzy/unmapped 혼재."""
    return [
        "전표번호", "posting_date", "GL코드",
        "차변", "Credit Amount", "메모",
        "담당자", "XYZ_UNKNOWN",
    ]
