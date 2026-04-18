"""ExcelExporter 단위 테스트.

테스트 그룹:
    - TestBasicExport: 정상 데이터로 시트 5~6개 생성, 시트명 확인
    - TestFilters: ExportFilter 적용 시 행 필터링 동작
    - TestMasking: mask_pii=True 시 작성자/승인자 해싱
    - TestStyling: WriteOnlyCell + 위험등급 fill 색상 적용
    - TestEmptyData: 빈 데이터/배치에서 graceful 처리
"""

from __future__ import annotations

from dataclasses import dataclass

import duckdb
import pandas as pd
import pytest
from openpyxl import load_workbook

from src.db.schema import initialize_schema
from src.export.excel_exporter import ExcelExporter
from src.export.models import ExportConfig, ExportFilter

BATCH = "TEST_BATCH_001"


@dataclass
class _StubPipelineResult:
    """PipelineResult의 최소 필드 stub — 테스트가 의존하는 속성만 모방."""

    data: pd.DataFrame
    results: list
    risk_summary: dict
    batch_id: str
    elapsed: float = 1.23
    load_result: object | None = None
    warnings: list = None  # type: ignore[assignment]


@pytest.fixture
def conn(tmp_path):
    """격리된 DuckDB + 스키마 초기화 + 샘플 데이터 적재."""
    db_path = tmp_path / "excel_test.duckdb"
    c = duckdb.connect(str(db_path))
    initialize_schema(c)

    # Why: schema_supplementary는 ExcelExporter가 사용하지 않음 → 생략 가능.
    _seed_sample_data(c)
    yield c
    c.close()


def _seed_sample_data(c: duckdb.DuckDBPyConnection) -> None:
    """5건 전표(High 2, Medium 1, Low 1, Normal 1) + Benford + flags."""
    rows = [
        ("D001", "C001", 2026, 1, "2026-01-15 10:00:00", "SA", "P2P",
         "alice", "manager1", 1000.0, 0.0, "4100", 0.85, "High", "B05,C08"),
        ("D002", "C001", 2026, 1, "2026-01-16 23:00:00", "SA", "O2C",
         "bob", "manager1", 0.0, 500.0, "1100", 0.72, "High", "C03"),
        ("D003", "C002", 2026, 2, "2026-02-10 11:00:00", "KR", "TRE",
         "alice", None, 2000.0, 0.0, "2100", 0.50, "Medium", "B09"),
        ("D004", "C002", 2026, 3, "2026-03-01 14:00:00", "DR", "R2R",
         "carol", "manager2", 0.0, 300.0, "5100", 0.25, "Low", "B08"),
        ("D005", "C001", 2026, 4, "2026-04-05 09:00:00", "SA", "P2P",
         "alice", "manager1", 100.0, 0.0, "1100", 0.05, "Normal", ""),
    ]
    for r in rows:
        c.execute(
            """
            INSERT INTO general_ledger
              (document_id, company_code, fiscal_year, fiscal_period,
               posting_date, document_type, business_process,
               created_by, approved_by, debit_amount, credit_amount,
               gl_account, anomaly_score, risk_level, flagged_rules,
               line_number, upload_batch_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            [*r, 1, BATCH],
        )
    # anomaly_flags: B05 + C08 + B07 (SoD)
    flags = [
        ("D001", "layer_b", "B05", 0.9),
        ("D001", "layer_c", "C08", 0.85),
        ("D002", "layer_c", "C03", 0.72),
        ("D003", "layer_b", "B09", 0.5),
        ("D003", "layer_b", "B07", 0.6),
    ]
    for f in flags:
        c.execute(
            """
            INSERT INTO anomaly_flags
              (upload_batch_id, document_id, line_number, track_name, rule_code, score)
            VALUES (?,?,?,?,?,?)
            """,
            [BATCH, f[0], 1, f[1], f[2], f[3]],
        )
    # benford
    c.execute(
        "INSERT INTO benford_summary (upload_batch_id, sample_size, mad, "
        "mad_conformity, chi2_statistic, chi2_p_value, is_conforming, confidence) "
        "VALUES (?,?,?,?,?,?,?,?)",
        [BATCH, 5, 0.012, "close", 1.5, 0.99, True, "low"],
    )
    for digit in range(1, 10):
        c.execute(
            "INSERT INTO benford_digits "
            "(upload_batch_id, digit, observed_freq, expected_freq, deviation) "
            "VALUES (?,?,?,?,?)",
            [BATCH, digit, 0.1, 0.1, 0.0],
        )


@pytest.fixture
def pipeline_result():
    df = pd.DataFrame({"document_id": ["D001"], "anomaly_score": [0.85]})
    return _StubPipelineResult(
        data=df,
        results=[],
        risk_summary={"High": 2, "Medium": 1, "Low": 1, "Normal": 1},
        batch_id=BATCH,
        warnings=[],
    )


# ── Tests ────────────────────────────────────────────────────


class TestBasicExport:
    def test_creates_file_with_six_sheets(self, conn, pipeline_result, tmp_path):
        out = tmp_path / "report.xlsx"
        ExcelExporter(conn).export(pipeline_result, out)
        assert out.exists() and out.stat().st_size > 0

        wb = load_workbook(out, read_only=True)
        names = wb.sheetnames
        # Why: include_raw_data=True 기본 → 6시트
        assert names == [
            "분석 요약", "이상 전표", "Benford 분석",
            "탐지 규칙 통계", "직무분리 분석", "원본 데이터",
        ]

    def test_skip_raw_data_yields_five_sheets(self, conn, pipeline_result, tmp_path):
        out = tmp_path / "report.xlsx"
        ExcelExporter(conn).export(
            pipeline_result, out, config=ExportConfig(include_raw_data=False)
        )
        wb = load_workbook(out, read_only=True)
        assert "원본 데이터" not in wb.sheetnames
        assert len(wb.sheetnames) == 5

    def test_summary_sheet_contains_disclaimer(self, conn, pipeline_result, tmp_path):
        out = tmp_path / "report.xlsx"
        ExcelExporter(conn).export(pipeline_result, out)

        wb = load_workbook(out, read_only=True)
        ws = wb["분석 요약"]
        all_text = "\n".join(
            str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
        )
        assert "감사 의견" in all_text  # 면책조항 키워드
        assert "데이터 분석" in all_text
        assert pipeline_result.batch_id in all_text


class TestFilters:
    def test_filter_by_company_code(self, conn, pipeline_result, tmp_path):
        out = tmp_path / "report.xlsx"
        ExcelExporter(conn).export(
            pipeline_result, out,
            filters=ExportFilter(company_codes=["C001"]),
        )
        wb = load_workbook(out, read_only=True)
        ws = wb["이상 전표"]
        rows = list(ws.iter_rows(values_only=True))
        # Why: C001 + risk!=Normal → D001, D002 두 건만 남아야 함
        document_ids = [r[0] for r in rows[1:]]  # 헤더 제외
        assert set(document_ids) == {"D001", "D002"}

    def test_filter_by_risk_level(self, conn, pipeline_result, tmp_path):
        out = tmp_path / "report.xlsx"
        ExcelExporter(conn).export(
            pipeline_result, out,
            filters=ExportFilter(risk_levels=["High"]),
        )
        wb = load_workbook(out, read_only=True)
        ws = wb["이상 전표"]
        rows = list(ws.iter_rows(values_only=True))
        document_ids = [r[0] for r in rows[1:]]
        assert set(document_ids) == {"D001", "D002"}


class TestMasking:
    def test_pii_masking_applies_to_anomalies_sheet(self, conn, pipeline_result, tmp_path):
        out = tmp_path / "report.xlsx"
        ExcelExporter(conn).export(
            pipeline_result, out, config=ExportConfig(mask_pii=True)
        )
        wb = load_workbook(out, read_only=True)
        ws = wb["이상 전표"]
        rows = list(ws.iter_rows(values_only=True))
        header = list(rows[0])
        created_idx = header.index("created_by")
        for row in rows[1:]:
            value = row[created_idx]
            # SHA-256 8자리 hex 또는 sentinel("--------")
            assert value is None or len(value) == 8


class TestStyling:
    def test_high_risk_row_has_red_fill(self, conn, pipeline_result, tmp_path):
        out = tmp_path / "report.xlsx"
        ExcelExporter(conn).export(pipeline_result, out)
        # write_only로 저장한 파일을 read_only=False로 다시 로드해야 fill 확인 가능
        wb = load_workbook(out)
        ws = wb["이상 전표"]
        rows = list(ws.iter_rows())
        header_cells = rows[0]
        risk_idx = next(
            i for i, c in enumerate(header_cells) if c.value == "risk_level"
        )
        # High 행 찾기
        for row in rows[1:]:
            if row[risk_idx].value == "High":
                fill_color = row[0].fill.fgColor.rgb
                assert "FFC7CE" in str(fill_color), fill_color
                return
        pytest.fail("High 위험 행을 찾지 못함")


class TestEmptyData:
    def test_unknown_batch_produces_empty_sheets(self, conn, tmp_path):
        empty_pr = _StubPipelineResult(
            data=pd.DataFrame(),
            results=[],
            risk_summary={},
            batch_id="NONEXISTENT_BATCH",
            warnings=[],
        )
        out = tmp_path / "empty.xlsx"
        # Why: 존재하지 않는 배치라도 graceful — 파일 자체는 생성되어야 함
        ExcelExporter(conn).export(empty_pr, out)
        assert out.exists()
        wb = load_workbook(out, read_only=True)
        assert "분석 요약" in wb.sheetnames
