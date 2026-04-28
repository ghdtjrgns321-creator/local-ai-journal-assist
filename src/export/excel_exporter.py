"""WU-24 Excel 데이터분석 보고서 생성기.

Why:
    감사인이 분석 결과를 자신의 감사조서에 첨부할 수 있도록 5~6시트로
    구조화된 Excel을 출력한다. ``감사조서`` 자체가 아닌 ``데이터 분석
    결과 보고서`` 위치임을 시트 1 면책조항에서 명시한다.

설계 결정:
    - openpyxl ``write_only=True`` 모드 (Raw Data 106K행 메모리 대응)
    - ``WriteOnlyCell`` + ``PatternFill`` 으로 조건부 서식 (사후 셀 접근 불가 회피)
    - DuckDB 직접 쿼리 + 파라미터 바인딩 (SQL Injection 방어)
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import TYPE_CHECKING, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill

from src.detection.constants import get_track_display_label
from src.export.audit_evidence import build_evidence_row
from src.export.masking import mask_dataframe
from src.export.models import (
    DETECTION_COLUMNS,
    DISCLAIMER,
    HEADER_COLUMNS,
    LINE_COLUMNS,
    RISK_FILL_COLORS,
    ExportConfig,
    ExportFilter,
)
from src.export.phase1_case_view import build_phase1_case_queue, summarize_phase1_case_result
from src.export.query_helper import build_where_clause, safe_query

if TYPE_CHECKING:
    import duckdb

    from src.pipeline import PipelineResult

logger = logging.getLogger(__name__)

# Why: PatternFill 객체는 한 번만 생성하여 모든 셀에서 재사용 (메모리/생성 비용).
_FILLS: dict[str, PatternFill] = {
    risk: PatternFill("solid", fgColor=color) for risk, color in RISK_FILL_COLORS.items()
}
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


class ExcelExporter:
    """DuckDB 기반 데이터분석 보고서 Excel 생성기."""

    def __init__(self, conn: duckdb.DuckDBPyConnection) -> None:
        self._conn = conn

    # ── public API ────────────────────────────────────────────
    def export(
        self,
        pipeline_result: PipelineResult,
        output_path: Path,
        filters: ExportFilter | None = None,
        config: ExportConfig | None = None,
    ) -> Path:
        """Excel 파일을 생성하고 경로를 반환한다."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(
            self.export_bytes(
                pipeline_result,
                filters=filters,
                config=config,
            )
        )
        return output_path

    def export_bytes(
        self,
        pipeline_result: PipelineResult,
        filters: ExportFilter | None = None,
        config: ExportConfig | None = None,
    ) -> bytes:
        filters = filters or ExportFilter()
        config = config or ExportConfig()

        wb = Workbook(write_only=True)
        where_sql, params = build_where_clause(filters, pipeline_result.batch_id)

        self._write_summary_sheet(wb, pipeline_result, config)
        self._write_anomalies_sheet(wb, pipeline_result, config, where_sql, params)
        self._write_benford_sheet(wb, pipeline_result.batch_id)
        self._write_rules_sheet(wb, pipeline_result.batch_id)
        self._write_sod_sheet(wb, pipeline_result.batch_id, where_sql, params)
        if config.include_raw_data:
            self._write_raw_data_sheet(wb, config, where_sql, params)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ── 시트별 ────────────────────────────────────────────────
    def _write_summary_sheet(
        self, wb: Workbook, pr: PipelineResult, config: ExportConfig
    ) -> None:
        """시트 1: 분석 요약 + 위험 분포 + 면책조항."""
        ws = wb.create_sheet("분석 요약")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 24

        ws.append([self._title_cell(ws, config.report_title)])
        ws.append([self._cell(ws, DISCLAIMER, italic=True)])
        ws.append([])

        ws.append([self._header_cell(ws, "지표"), self._header_cell(ws, "값")])
        kpi_rows = [
            ("배치 ID", pr.batch_id),
            ("총 행 수", len(pr.data)),
            ("분석 소요(초)", round(pr.elapsed, 2)),
            ("탐지 트랙 수", len(pr.results)),
        ]
        if config.analyst_name:
            kpi_rows.append(("분석자", config.analyst_name))
        for label, value in kpi_rows:
            ws.append([label, value])

        ws.append([])
        ws.append([self._header_cell(ws, "위험 등급"), self._header_cell(ws, "건수")])
        # Why: risk_summary가 비어 있어도 표 머리만 두고 빈 시트가 되지 않게 한다.
        for risk in ("High", "Medium", "Low", "Normal"):
            count = pr.risk_summary.get(risk, 0)
            cell_a = self._cell(ws, risk, fill=_FILLS.get(risk))
            ws.append([cell_a, count])

        phase1_summary = summarize_phase1_case_result(pr)
        if config.include_phase1_cases and phase1_summary.get("available"):
            ws.append([])
            ws.append([self._header_cell(ws, "PHASE1 Case Queue"), self._header_cell(ws, "값")])
            ws.append(["Case 수", phase1_summary["case_count"]])
            ws.append(["Top Themes", ", ".join(phase1_summary.get("top_theme_labels", []))])
            top_cases = build_phase1_case_queue(pr, top_n=5)
            if top_cases:
                ws.append([])
                ws.append(
                    [
                        self._header_cell(ws, "Case ID"),
                        self._header_cell(ws, "Theme"),
                        self._header_cell(ws, "Band"),
                        self._header_cell(ws, "Amount"),
                        self._header_cell(ws, "Risk Narrative"),
                        self._header_cell(ws, "Recommended Actions"),
                    ]
                )
                for case in top_cases:
                    ws.append(
                        [
                            case["case_id"],
                            case["primary_theme_label"],
                            case["priority_band"],
                            case["total_amount"],
                            case.get("risk_narrative") or case["representative_explanation"],
                            "; ".join(case.get("recommended_audit_actions", [])),
                        ]
                    )

    def _write_anomalies_sheet(
        self,
        wb: Workbook,
        pr: PipelineResult,
        config: ExportConfig,
        where_sql: str,
        params: list[Any],
    ) -> None:
        """시트 2: risk_level != Normal 전표 + narrative 컬럼."""
        ws = wb.create_sheet("이상 전표")
        ws.freeze_panes = "A2"

        sql = f"""
            SELECT document_id, company_code, posting_date, document_type,
                   business_process, gl_account, debit_amount, credit_amount,
                   created_by, approved_by, anomaly_score, risk_level, flagged_rules
            FROM general_ledger
            WHERE risk_level IS NOT NULL AND risk_level <> 'Normal'
              {where_sql}
            ORDER BY anomaly_score DESC NULLS LAST
        """
        df = self._safe_query(sql, params)
        if config.mask_pii and not df.empty:
            df = mask_dataframe(df)

        # Why: AuditEvidence 문구를 추가 컬럼으로 노출 — 감사인이 한국어로 즉시 이해.
        if not df.empty:
            df["분석 증거 문구"] = df.apply(
                lambda row: build_evidence_row(row).narrative, axis=1
            )

        headers = list(df.columns)
        ws.append([self._header_cell(ws, h) for h in headers])
        for _, row in df.iterrows():
            risk = str(row.get("risk_level", ""))
            ws.append([self._cell(ws, v, fill=_FILLS.get(risk)) for v in row])

    def _write_benford_sheet(self, wb: Workbook, batch_id: str) -> None:
        """시트 3: Benford 요약 + 자릿수 분포."""
        ws = wb.create_sheet("Benford 분석")

        summary = self._safe_query(
            "SELECT * FROM benford_summary WHERE upload_batch_id = ?", [batch_id]
        )
        ws.append([self._title_cell(ws, "Benford 적합도 요약")])
        if summary.empty:
            ws.append([self._cell(ws, "데이터 없음 (Benford 분석 미수행)")])
        else:
            ws.append([self._header_cell(ws, c) for c in summary.columns])
            for _, row in summary.iterrows():
                ws.append([self._cell(ws, v) for v in row])

        ws.append([])
        digits = self._safe_query(
            "SELECT digit, observed_freq, expected_freq, deviation "
            "FROM benford_digits WHERE upload_batch_id = ? ORDER BY digit",
            [batch_id],
        )
        ws.append([self._title_cell(ws, "자릿수별 분포")])
        if digits.empty:
            ws.append([self._cell(ws, "데이터 없음")])
        else:
            ws.append([self._header_cell(ws, c) for c in digits.columns])
            for _, row in digits.iterrows():
                ws.append([self._cell(ws, v) for v in row])

    def _write_rules_sheet(self, wb: Workbook, batch_id: str) -> None:
        """시트 4: 탐지 규칙별 통계."""
        ws = wb.create_sheet("탐지 규칙 통계")
        sql = """
            SELECT track_name, rule_code,
                   COUNT(*) AS 탐지건수,
                   ROUND(AVG(score), 4) AS 평균점수,
                   ROUND(MAX(score), 4) AS 최고점수
            FROM anomaly_flags
            WHERE upload_batch_id = ?
            GROUP BY track_name, rule_code
            ORDER BY track_name, rule_code
        """
        df = self._safe_query(sql, [batch_id])
        if {"track_name", "rule_code"}.issubset(df.columns):
            df.insert(
                0,
                "rule_group",
                [
                    get_track_display_label(track_name, rule_code)
                    for track_name, rule_code in zip(
                        df["track_name"],
                        df["rule_code"],
                        strict=True,
                    )
                ],
            )
            df = df.drop(columns=["track_name"])
        if df.empty:
            ws.append([self._cell(ws, "탐지된 규칙 없음")])
            return
        ws.append([self._header_cell(ws, c) for c in df.columns])
        for _, row in df.iterrows():
            ws.append([self._cell(ws, v) for v in row])

    def _write_sod_sheet(
        self, wb: Workbook, batch_id: str, where_sql: str, params: list[Any]
    ) -> None:
        """시트 5: 직무분리(SoD) 위반 상세 — L1-05/L1-06/L1-07 룰.

        Why:
            CTE로 일반 원장 필터를 먼저 적용하여 ``_build_where_clause``의
            컬럼명을 그대로 활용 (테이블 alias 충돌 회피).
        """
        ws = wb.create_sheet("직무분리 분석")
        sql = f"""
            WITH filtered_gl AS (
                SELECT document_id, company_code, posting_date,
                       created_by, approved_by, user_persona,
                       business_process, upload_batch_id
                FROM general_ledger
                WHERE 1=1 {where_sql}
            )
            SELECT gl.document_id, gl.company_code, gl.posting_date,
                   gl.created_by, gl.approved_by, gl.user_persona,
                   gl.business_process, af.rule_code, af.score
            FROM anomaly_flags af
            JOIN filtered_gl gl USING (document_id, upload_batch_id)
            WHERE af.upload_batch_id = ?
              AND af.rule_code IN ('L1-05', 'L1-06', 'L1-07')
            ORDER BY af.score DESC
        """
        df = self._safe_query(sql, [*params, batch_id])
        if df.empty:
            ws.append([self._cell(ws, "직무분리 위반 없음")])
            return
        ws.append([self._header_cell(ws, c) for c in df.columns])
        for _, row in df.iterrows():
            ws.append([self._cell(ws, v) for v in row])

    def _write_raw_data_sheet(
        self,
        wb: Workbook,
        config: ExportConfig,
        where_sql: str,
        params: list[Any],
    ) -> None:
        """시트 6: document_id 단위 집계 원본 데이터."""
        ws = wb.create_sheet("원본 데이터")
        ws.freeze_panes = "A2"

        sql = f"""
            SELECT document_id, company_code, fiscal_year, fiscal_period,
                   MIN(posting_date) AS posting_date,
                   MAX(document_type) AS document_type,
                   MAX(source) AS source,
                   MAX(business_process) AS business_process,
                   MAX(header_text) AS header_text,
                   MAX(created_by) AS created_by,
                   MAX(approved_by) AS approved_by,
                   COUNT(line_number) AS 라인수,
                   ROUND(SUM(debit_amount), 2) AS 차변합계,
                   ROUND(SUM(credit_amount), 2) AS 대변합계,
                   STRING_AGG(DISTINCT gl_account, ', ') AS 사용계정,
                   MAX(anomaly_score) AS anomaly_score,
                   MAX(risk_level) AS risk_level,
                   STRING_AGG(DISTINCT flagged_rules, ', ') AS flagged_rules
            FROM general_ledger
            WHERE 1=1 {where_sql}
            GROUP BY document_id, company_code, fiscal_year, fiscal_period
            ORDER BY anomaly_score DESC NULLS LAST
        """
        df = self._safe_query(sql, params)
        if config.mask_pii and not df.empty:
            df = mask_dataframe(df)

        if df.empty:
            ws.append([self._cell(ws, "원본 데이터 없음")])
            return
        ws.append([self._header_cell(ws, c) for c in df.columns])
        for _, row in df.iterrows():
            risk = str(row.get("risk_level", ""))
            ws.append([self._cell(ws, v, fill=_FILLS.get(risk)) for v in row])

    # ── helpers ───────────────────────────────────────────────
    def _safe_query(self, sql: str, params: list[Any]) -> pd.DataFrame:
        """query_helper.safe_query 위임 (인스턴스 메서드 호환 유지)."""
        return safe_query(self._conn, sql, params)

    def _cell(
        self,
        ws: Any,
        value: Any,
        fill: PatternFill | None = None,
        italic: bool = False,
    ) -> WriteOnlyCell:
        """write_only 모드용 셀 팩토리. 서식 직접 적용."""
        cell = WriteOnlyCell(ws, value=self._normalize_value(value))
        if fill is not None:
            cell.fill = fill
        if italic:
            cell.font = Font(italic=True)
        return cell

    def _header_cell(self, ws: Any, value: Any) -> WriteOnlyCell:
        cell = WriteOnlyCell(ws, value=self._normalize_value(value))
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        return cell

    def _title_cell(self, ws: Any, value: str) -> WriteOnlyCell:
        cell = WriteOnlyCell(ws, value=value)
        cell.font = Font(bold=True, size=14)
        return cell

    @staticmethod
    def _normalize_value(value: Any) -> Any:
        """openpyxl이 직접 처리할 수 없는 값(NaN/NaT/Timestamp 등)을 변환.

        Why:
            float NaN만 잡으면 ``pd.NaT``(timestamp 결측)가 누락되어
            openpyxl 직렬화 시 경고/오류가 발생한다. ``pd.isna``를
            먼저 시도하되, 비스칼라 입력에 대한 TypeError는 안전하게 무시.
        """
        if value is None:
            return None
        try:
            if pd.isna(value):
                return None
        except (TypeError, ValueError):
            # Why: 리스트/배열 등 비스칼라는 isna가 ambiguous 에러를 던질 수 있음.
            pass
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        return value


__all__ = [
    "ExcelExporter",
    "DETECTION_COLUMNS",
    "HEADER_COLUMNS",
    "LINE_COLUMNS",
]
