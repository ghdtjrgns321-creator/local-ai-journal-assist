"""엑셀 리더 — xlsx/xls/xlsb 파일을 ReadResult로 변환.

병합셀 처리를 위해 read_only=False를 사용한다.
file_validator의 100MB 제한이 메모리 안전장치 역할을 한다.
"""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd

from src.ingest.models import ReadResult

logger = logging.getLogger(__name__)


# --- xlsx (openpyxl) ---

def _unmerge_and_fill(ws) -> None:
    """워크시트의 병합셀을 해제하고 좌상단 값을 모든 셀에 복제한다.

    openpyxl은 unmerge 시 좌상단 값만 남기고 나머지를 None으로 만든다.
    따라서 unmerge 전에 좌상단 값을 저장해두고, unmerge 후 전체에 채운다.
    """
    # list()로 복사 — 순회 중 ranges가 변경되므로
    for merged_range in list(ws.merged_cells.ranges):
        top_left_value = ws.cell(
            row=merged_range.min_row,
            column=merged_range.min_col,
        ).value

        ws.unmerge_cells(str(merged_range))

        # 해제된 범위의 모든 셀에 좌상단 값 복제
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                ws.cell(row=row, column=col, value=top_left_value)


def _read_xlsx(path: Path) -> ReadResult:
    """openpyxl로 .xlsx 파일을 읽는다. 병합셀 해제 + 값 복제."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        sheet_names = wb.sheetnames
        active_name = wb.active.title if wb.active else sheet_names[0]
        raw_data: dict[str, pd.DataFrame] = {}

        for name in sheet_names:
            ws = wb[name]
            _unmerge_and_fill(ws)
            # ws.values는 제너레이터 — list로 변환 후 DataFrame 생성
            rows = list(ws.values)
            if rows:
                raw_data[name] = pd.DataFrame(rows)
            else:
                raw_data[name] = pd.DataFrame()
    finally:
        wb.close()

    return ReadResult(
        sheets=sheet_names,
        active_sheet=active_name,
        raw_data=raw_data,
        encoding=None,
        source_format="xlsx",
    )


# --- xls (xlrd) ---

def _read_xls(path: Path) -> ReadResult:
    """xlrd로 .xls 파일을 읽는다. 병합셀 해제 + 값 복제."""
    import xlrd

    wb = xlrd.open_workbook(str(path), formatting_info=True)
    try:
        sheet_names = wb.sheet_names()
        raw_data: dict[str, pd.DataFrame] = {}

        for name in sheet_names:
            sheet = wb.sheet_by_name(name)

            # 2D 리스트로 변환
            rows = []
            for row_idx in range(sheet.nrows):
                rows.append(
                    [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
                )

            # 병합셀 값 복제 — xlrd의 merged_cells는 (rlo, rhi, clo, chi) 튜플
            for rlo, rhi, clo, chi in sheet.merged_cells:
                top_left = rows[rlo][clo] if rows else None
                for r in range(rlo, rhi):
                    for c in range(clo, chi):
                        rows[r][c] = top_left

            if rows:
                raw_data[name] = pd.DataFrame(rows)
            else:
                raw_data[name] = pd.DataFrame()
    finally:
        wb.release_resources()

    return ReadResult(
        sheets=sheet_names,
        # xlrd는 active_sheet 개념 없음 → 첫 시트를 기본값으로 설정
        active_sheet=sheet_names[0] if sheet_names else "",
        raw_data=raw_data,
        encoding=None,
        source_format="xls",
    )


# --- xlsb (pyxlsb) ---

def _read_xlsb(path: Path) -> ReadResult:
    """pyxlsb로 .xlsb 파일을 읽는다. 병합셀 정보 없음 → warning."""
    import pyxlsb

    with pyxlsb.open_workbook(str(path)) as wb:
        sheet_names = wb.sheets
        raw_data: dict[str, pd.DataFrame] = {}

        logger.warning(
            "xlsb 포맷은 병합셀 정보를 제공하지 않습니다. "
            "병합셀이 있으면 데이터가 누락될 수 있습니다: %s",
            path.name,
        )

        for name in sheet_names:
            rows = []
            with wb.get_sheet(name) as sheet:
                for row in sheet.rows():
                    rows.append([cell.v for cell in row])

            if rows:
                raw_data[name] = pd.DataFrame(rows)
            else:
                raw_data[name] = pd.DataFrame()

    return ReadResult(
        sheets=sheet_names,
        # pyxlsb는 active_sheet 개념 없음 → 첫 시트를 기본값으로 설정
        active_sheet=sheet_names[0] if sheet_names else "",
        raw_data=raw_data,
        encoding=None,
        source_format="xlsb",
    )


# --- 퍼블릭 API ---

# 확장자 → 내부 리더 매핑
_EXCEL_READERS = {
    ".xlsx": _read_xlsx,
    ".xls": _read_xls,
    ".xlsb": _read_xlsb,
}


def read_excel(path: Path) -> ReadResult:
    """엑셀 파일(xlsx/xls/xlsb)을 읽어 ReadResult를 반환한다.

    Raises:
        ValueError: 지원하지 않는 엑셀 확장자.
        OSError: 파일 읽기 실패 시.
    """
    ext = path.suffix.lower()
    reader = _EXCEL_READERS.get(ext)

    if reader is None:
        msg = f"지원하지 않는 엑셀 확장자입니다: {ext}"
        raise ValueError(msg)

    return reader(path)
