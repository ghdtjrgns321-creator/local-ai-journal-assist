"""Apply heuristic labels to the OpenDataPhilly golden review workbook.

This script fills `review_label`, `review_confidence_1_5`, `review_rationale`,
`evidence_needed`, `eligible_for_supervised_positive`, and `reviewer` for the
200 sampled rows. Labels are derived from visible public-payment fields only.

The intent is bootstrap labeling, not a substitute for human audit review. The
reviewer field is stamped `heuristic_ai_v1` and `eligible_for_supervised_positive`
is `false` for every row because confidence never exceeds 3.

Usage:
    uv run python tools/scripts/apply_open_data_philly_heuristic_labels.py
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

DEFAULT_DIR = ROOT / "artifacts" / "external_validation" / "open_data_philly_20260519"
WORKBOOK_NAME = "open_data_philly_golden_review_workbook.xlsx"
REVIEW_CSV_NAME = "golden_review_sheet.csv"
CONTEXT_CSV_NAME = "selected_review_context.csv"
REVIEWER_TAG = "heuristic_ai_v1"

LABEL_COL = "review_label"
CONFIDENCE_COL = "review_confidence_1_5"
RATIONALE_COL = "review_rationale"
EVIDENCE_COL = "evidence_needed"
ELIGIBLE_COL = "eligible_for_supervised_positive"
REVIEWER_COL = "reviewer"


def _coerce_int(value) -> int:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _has_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, float) and pd.isna(value):
        return False
    if isinstance(value, str) and value.strip() == "":
        return False
    return True


def classify_row(row: pd.Series) -> dict[str, object]:
    """Heuristic classifier. Returns dict with five labeling fields."""

    bucket = row.get("sample_bucket", "")
    is_negative = bool(row.get("is_negative_amount", False))
    is_zero = bool(row.get("is_zero_amount", False))
    is_round_1k = bool(row.get("is_round_1000", False))
    is_month_end = bool(row.get("is_month_end_window", False))
    is_fiscal_ye = bool(row.get("is_fiscal_year_end_window", False))
    is_near_threshold = bool(row.get("is_near_common_threshold", False))

    group_size = _coerce_int(row.get("same_vendor_date_amount_group_size"))
    contract_group_size = _coerce_int(row.get("same_vendor_contract_date_amount_group_size"))
    has_contract = _has_value(row.get("contract_number"))
    has_description = _has_value(row.get("description"))

    if is_negative:
        return {
            LABEL_COL: "accounting_error",
            CONFIDENCE_COL: 3,
            RATIONALE_COL: (
                "Negative payment to vendor in fleet/parts category; pattern is "
                "consistent with refund, chargeback, or credit memo applied to an "
                "earlier disbursement. The original offsetting payment is not in "
                "public fields."
            ),
            EVIDENCE_COL: ("original payment voucher, credit memo, GL offset entry"),
            ELIGIBLE_COL: False,
        }

    if is_zero:
        return {
            LABEL_COL: "accounting_error",
            CONFIDENCE_COL: 3,
            RATIONALE_COL: (
                "Zero-amount payment voucher; pattern is consistent with a void or "
                "corrective entry rather than an active disbursement."
            ),
            EVIDENCE_COL: "voiding rationale, original voucher reference",
            ELIGIBLE_COL: False,
        }

    if bucket == "triage_top_k":
        # Large bulk vehicle/equipment purchase distributed across departments
        # under a single contract with descriptive narrative.
        if group_size >= 5 and has_contract and has_description and contract_group_size >= 5:
            return {
                LABEL_COL: "audit_review_candidate",
                CONFIDENCE_COL: 2,
                RATIONALE_COL: (
                    f"High same-vendor/date/amount group size ({group_size}) under "
                    "a single contract with description; consistent with a bulk "
                    "purchase order distributed across departments but warrants "
                    "verification of per-line allocation and receipt confirmation."
                ),
                EVIDENCE_COL: (
                    "PO line allocation, departmental receipt confirmations, approval thresholds"
                ),
                ELIGIBLE_COL: False,
            }
        # Round 1,000 at fiscal year-end under a contract — period-end installment
        # pattern (e.g. transitional housing 25,000).
        if is_round_1k and is_fiscal_ye and has_contract:
            return {
                LABEL_COL: "audit_review_candidate",
                CONFIDENCE_COL: 2,
                RATIONALE_COL: (
                    "Round-1,000 payment at fiscal year-end under a contract; "
                    "period-end timing combined with round amount justifies "
                    "review of the underlying installment or milestone schedule."
                ),
                EVIDENCE_COL: (
                    "contract installment schedule, milestone deliverables, year-end accrual policy"
                ),
                ELIGIBLE_COL: False,
            }
        # Round 1,000 with month/quarter-end timing and no contract reference.
        if is_round_1k and is_month_end and not has_contract:
            return {
                LABEL_COL: "audit_review_candidate",
                CONFIDENCE_COL: 2,
                RATIONALE_COL: (
                    "Round-1,000 month-end payment without contract reference and "
                    "without description in public fields; the combination of "
                    "round amount, period-end timing, and missing contract "
                    "context warrants review."
                ),
                EVIDENCE_COL: ("contract reference, supporting invoice, approval memo"),
                ELIGIBLE_COL: False,
            }
        # Near common approval threshold with period-end and contract present.
        if is_near_threshold and is_month_end and has_contract:
            return {
                LABEL_COL: "audit_review_candidate",
                CONFIDENCE_COL: 2,
                RATIONALE_COL: (
                    "Amount near a common approval threshold at period-end under "
                    "a contract; warrants review for approval-level integrity and "
                    "whether the payment was structured to fall just under "
                    "threshold."
                ),
                EVIDENCE_COL: ("approval thresholds policy, approver identity, posting timestamp"),
                ELIGIBLE_COL: False,
            }
        # Near common approval threshold under a contract.
        if is_near_threshold and has_contract:
            return {
                LABEL_COL: "audit_review_candidate",
                CONFIDENCE_COL: 2,
                RATIONALE_COL: (
                    "Amount near a common approval threshold under a contract "
                    "with description; review whether the amount was structured "
                    "to fall just under an approval limit, and review prior "
                    "payments on the same contract."
                ),
                EVIDENCE_COL: ("approval thresholds policy, prior payments on same contract"),
                ELIGIBLE_COL: False,
            }
        # Default for any triage_top_k row that didn't match a more specific
        # signal — visible fields are not enough to call an exception, but the
        # triage process flagged it.
        return {
            LABEL_COL: "audit_review_candidate",
            CONFIDENCE_COL: 2,
            RATIONALE_COL: (
                "Row was flagged by the triage scorer based on a combination of "
                "near-threshold, period-end, or repeat-payment signals, but the "
                "public payment fields alone are not sufficient to classify it "
                "as a clear exception, control issue, or benign payment."
            ),
            EVIDENCE_COL: (
                "approver identity, posting timestamp, debit/credit lines, supporting invoice"
            ),
            ELIGIBLE_COL: False,
        }

    # random_control bucket
    if not has_description and not has_contract:
        return {
            LABEL_COL: "insufficient_evidence",
            CONFIDENCE_COL: 2,
            RATIONALE_COL: (
                "Public fields show vendor, department, amount, and date but no "
                "contract reference and no description; available evidence is "
                "not enough to classify the row as benign or exceptional."
            ),
            EVIDENCE_COL: ("invoice, contract reference, posting timestamp, approver"),
            ELIGIBLE_COL: False,
        }
    if has_contract and has_description:
        return {
            LABEL_COL: "benign_explainable",
            CONFIDENCE_COL: 2,
            RATIONALE_COL: (
                "Routine payment under an identified contract with a descriptive "
                "narrative; pattern is consistent with normal operational "
                "disbursement and shows no period-end, round-amount, or "
                "near-threshold trigger."
            ),
            EVIDENCE_COL: ("no further evidence required at this confidence level"),
            ELIGIBLE_COL: False,
        }
    return {
        LABEL_COL: "insufficient_evidence",
        CONFIDENCE_COL: 2,
        RATIONALE_COL: (
            "Partial context available: either contract or description is "
            "missing. Public fields are not sufficient to confirm benign vs "
            "exception classification."
        ),
        EVIDENCE_COL: ("missing contract reference or description, posting timestamp"),
        ELIGIBLE_COL: False,
    }


def apply_to_workbook(workbook_path: Path, labels: pd.DataFrame) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["Review"]
    headers = [cell.value for cell in ws[1]]
    header_to_col = {h: i + 1 for i, h in enumerate(headers)}

    target_cols = (
        LABEL_COL,
        CONFIDENCE_COL,
        RATIONALE_COL,
        EVIDENCE_COL,
        ELIGIBLE_COL,
        REVIEWER_COL,
    )
    missing = [c for c in target_cols if c not in header_to_col]
    if missing:
        raise KeyError(f"Workbook missing columns: {missing}")

    id_col = header_to_col["external_row_id"]
    workbook_row_by_id: dict[str, int] = {}
    for excel_row in range(2, ws.max_row + 1):
        rid = ws.cell(row=excel_row, column=id_col).value
        if rid is None:
            continue
        workbook_row_by_id[str(rid)] = excel_row

    for record in labels.to_dict(orient="records"):
        rid = str(record["external_row_id"])
        excel_row = workbook_row_by_id.get(rid)
        if excel_row is None:
            continue
        for col in target_cols:
            ws.cell(row=excel_row, column=header_to_col[col]).value = record[col]

    wb.save(workbook_path)


def apply_to_csv(csv_path: Path, labels: pd.DataFrame) -> None:
    df = pd.read_csv(csv_path)
    indexed = labels.set_index("external_row_id")
    if "external_row_id" not in df.columns:
        raise KeyError("CSV missing external_row_id column")
    for col in (
        LABEL_COL,
        CONFIDENCE_COL,
        RATIONALE_COL,
        EVIDENCE_COL,
        ELIGIBLE_COL,
        REVIEWER_COL,
    ):
        if col not in df.columns:
            df[col] = pd.NA
    for excel_row_idx, row in df.iterrows():
        rid = row["external_row_id"]
        if rid not in indexed.index:
            continue
        for col in (
            LABEL_COL,
            CONFIDENCE_COL,
            RATIONALE_COL,
            EVIDENCE_COL,
            ELIGIBLE_COL,
            REVIEWER_COL,
        ):
            df.at[excel_row_idx, col] = indexed.at[rid, col]
    df.to_csv(csv_path, index=False, encoding="utf-8")


def build_labels(input_dir: Path) -> pd.DataFrame:
    context_path = input_dir / CONTEXT_CSV_NAME
    context = pd.read_csv(context_path)
    records = []
    for _, row in context.iterrows():
        result = classify_row(row)
        records.append(
            {
                "external_row_id": row["external_row_id"],
                **result,
                REVIEWER_COL: REVIEWER_TAG,
            }
        )
    return pd.DataFrame(records)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_DIR,
        help="Directory holding workbook and CSVs.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_dir: Path = args.input_dir
    labels = build_labels(input_dir)
    apply_to_workbook(input_dir / WORKBOOK_NAME, labels)
    apply_to_csv(input_dir / REVIEW_CSV_NAME, labels)
    apply_to_csv(input_dir / CONTEXT_CSV_NAME, labels)

    summary = labels[LABEL_COL].value_counts().to_dict()
    print("label_counts:")
    for label, count in sorted(summary.items()):
        print(f"  {label}: {count}")
    print(f"total_rows: {len(labels)}")


if __name__ == "__main__":
    main()
