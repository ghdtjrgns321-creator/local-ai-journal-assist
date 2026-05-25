"""Export an Excel reviewer workbook for the OpenDataPhilly golden packet.

The workbook is a convenience wrapper around the CSV artifacts. It does not add
labels or train any model. It provides dropdown validation for the human review
fields and includes context sheets for duplicate clusters and vendor history.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

LABELS = (
    "confirmed_exception",
    "control_issue",
    "accounting_error",
    "audit_review_candidate",
    "benign_explainable",
    "insufficient_evidence",
)

BOOLEAN_VALUES = ("true", "false")

REVIEW_COLUMNS = (
    "review_label",
    "review_confidence_1_5",
    "review_rationale",
    "evidence_needed",
    "eligible_for_supervised_positive",
)

SHEETS = {
    "Review": "selected_review_context.csv",
    "DuplicateContext": "duplicate_cluster_context.csv",
    "VendorContext": "vendor_context.csv",
    "ValidationStatus": "golden_review_validation.csv",
}


def _read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path, dtype=str, keep_default_na=False)


def _write_instructions(writer: pd.ExcelWriter) -> None:
    rows = [
        {
            "topic": "Purpose",
            "instruction": (
                "Classify sampled public payment rows for a golden review set. "
                "Do not treat this as fraud determination."
            ),
        },
        {
            "topic": "Main sheet",
            "instruction": "Fill only the review_* and eligible_for_supervised_positive columns.",
        },
        {
            "topic": "Labels",
            "instruction": ", ".join(LABELS),
        },
        {
            "topic": "Positive eligibility",
            "instruction": (
                "Use true only for confirmed_exception/control_issue/accounting_error "
                "with confidence >= 4 and specific rationale."
            ),
        },
        {
            "topic": "Guardrail",
            "instruction": (
                "Duplicate group size, amount size, and check_date timing are context, "
                "not labels by themselves."
            ),
        },
        {
            "topic": "Validation",
            "instruction": (
                "After editing the workbook, export/update golden_review_sheet.csv and "
                "run tools/scripts/validate_golden_review_sheet.py."
            ),
        },
    ]
    pd.DataFrame(rows).to_excel(writer, sheet_name="Instructions", index=False)


def _autosize_and_style(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    review_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_cells in ws.columns:
            header = str(col_cells[0].value or "")
            max_len = max(len(str(cell.value or "")) for cell in col_cells[:200])
            width = min(max(max_len + 2, len(header) + 2), 48)
            ws.column_dimensions[col_cells[0].column_letter].width = width
            if header in REVIEW_COLUMNS:
                for cell in col_cells:
                    cell.fill = review_fill
        if ws.max_row > 1:
            ws.sheet_view.zoomScale = 90

    if "Review" in wb.sheetnames:
        ws = wb["Review"]
        headers = {str(cell.value): cell.column_letter for cell in ws[1]}
        label_col = headers.get("review_label")
        confidence_col = headers.get("review_confidence_1_5")
        eligible_col = headers.get("eligible_for_supervised_positive")
        if label_col:
            dv = DataValidation(type="list", formula1=f'"{",".join(LABELS)}"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{label_col}2:{label_col}{ws.max_row}")
        if confidence_col:
            dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5")
            ws.add_data_validation(dv)
            dv.add(f"{confidence_col}2:{confidence_col}{ws.max_row}")
        if eligible_col:
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(BOOLEAN_VALUES)}"',
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            dv.add(f"{eligible_col}2:{eligible_col}{ws.max_row}")

    wb.save(path)


def export_workbook(input_dir: Path, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_instructions(writer)
        for sheet_name, filename in SHEETS.items():
            if sheet_name == "ValidationStatus":
                validation_json = input_dir / "golden_review_validation.json"
                if validation_json.exists():
                    df = pd.read_json(validation_json, typ="series").reset_index()
                    df.columns = ["field", "value"]
                else:
                    df = pd.DataFrame()
            else:
                df = _read_csv(input_dir / filename)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    _autosize_and_style(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input-dir",
        required=True,
        help="Directory containing selected_review_context.csv and related artifacts.",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Path to output .xlsx workbook.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    export_workbook(Path(args.input_dir), Path(args.output))
    print(f"wrote {args.output}")


if __name__ == "__main__":
    main()
